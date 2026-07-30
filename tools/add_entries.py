#!/usr/bin/env python3
"""Write a NEW dated version of the vocabulary workbook with rows added.

Never edits a workbook in place. Reads the newest `YYYYMMDD
Tibetan_Vocabulary_Dictionary.xlsx` in the repo root, applies the requested
additions, and saves today's version alongside it. The previous versions stay
as the archive.

    python3 tools/add_entries.py --entries new_rows.json \
        --source-code NGO7 --source-name "Kagyü Ngöndro 7 – Guru Yoga"

Requires openpyxl. Used by the /updatedata skill; also fine to run by hand.

TWO KINDS OF ADDITION — this is the rule that keeps the sheet one-row-per-word:

  "new"    a word not yet in the sheet          -> appended as a new row
  "append" a word already in the sheet, met     -> its EXISTING row gains a
           again in this source                    source in column R and a
                                                   lesson in column S

Never add a second row for a word that already has one. Look it up by its exact
Tibetan (column B) and append instead.

ENTRIES FILE — a JSON list. For a new word, keys are the short column names
below; `tibetan` and `english` are required, everything else defaults to "".

    [
      {"op": "new",
       "tibetan": "…", "romanization": "…", "wylie": "…", "english": "…",
       "category": "noun", "verbStems": "", "volitionality": "", "register": "",
       "sanskrit": "", "exampleTibetan": "…", "exampleEnglish": "…",
       "notes": "…",
       "related":      [["…","gloss"], …],   // pairs; written as "a / b" + "x / y"
       "falseFriends": [["…","gloss"], …],
       "lesson": "Ngöndro 7 v.7a",
       "pronCheck": "derived – reviewed"},

      {"op": "append", "tibetan": "སེམས་", "lesson": "Ngöndro 7 v.7c"}
    ]

The paired columns are written from pair lists, so the '/' item separator and
the ';' sense separator can never be got the wrong way round by hand. A gloss
containing '/' is rejected — it would silently become two items on re-import.
"""

import argparse
import copy
import datetime
import json
import pathlib
import re
import sys

WORKBOOK_GLOB = "20* Tibetan_Vocabulary_Dictionary.xlsx"
# "20260729 …" or, for a second version on the same day, "20260729 1930 …".
WORKBOOK_RE = re.compile(r"^(\d{8})(?: (\d{4}))? Tibetan_Vocabulary_Dictionary\.xlsx$")
SHEET = "Dictionary"

# Short name -> column letter. The full headers live in row 1 of the sheet.
FIELDS = {
    "id": "A", "tibetan": "B", "romanization": "C", "wylie": "D",
    "english": "E", "category": "F", "verbStems": "G", "volitionality": "H",
    "register": "I", "sanskrit": "J", "exampleTibetan": "K",
    "exampleEnglish": "L", "notes": "M", "relatedTibetan": "N",
    "relatedEnglish": "O", "falseFriendsTibetan": "P", "falseFriendsEnglish": "Q",
    "source": "R", "lesson": "S", "mastery": "T", "dateLearned": "U",
    "dateAdded": "V", "pronCheck": "W",
}
PAIRED = {"related": ("relatedTibetan", "relatedEnglish"),
          "falseFriends": ("falseFriendsTibetan", "falseFriendsEnglish")}

ITEM_SEP = " / "
SOURCE_SEP = "; "
LESSON_SEP = " | "
VALID_PRON = {"from source", "from slides", "derived – reviewed"}

# The layout the About sheet documents. Excel saves whatever the last human
# scroll position was, so this is normalised rather than inherited.
FREEZE_PANES = "F2"


def register_source(root, code, name, short, dry_run):
    """Record a new code in tools/source_names.json so the importer knows it.

    Without this the importer refuses the workbook, and the Look up scope chip
    falls back to a label that may collide with another source's.
    """
    path = root / "tools" / "source_names.json"
    with path.open(encoding="utf-8") as fh:
        doc = json.load(fh)
    if code in doc["sources"]:
        return False
    if not short:
        # Text before the first dash or bracket, the same reduction the app used
        # to do — but checked for collisions here, where it can still be fixed.
        short = re.split(r"\s[–\-(]", name)[0].strip()
    taken = {v.get("short") for v in doc["sources"].values()}
    if short in taken:
        sys.exit(f"short label {short!r} for {code} collides with an existing "
                 "source. Pass --source-short with something distinct.")
    doc["sources"][code] = {"name": name, "short": short}
    if not dry_run:
        with path.open("w", encoding="utf-8") as fh:
            json.dump(doc, fh, ensure_ascii=False, indent=2)
            fh.write("\n")
    return short


def workbook_key(path):
    """Sort key from the filename's date and optional time.

    Plain lexicographic sorting gets this wrong: '20260729 1930 …' would sort
    before '20260729 …' because '1' < 'T'.
    """
    m = WORKBOOK_RE.match(path.name)
    return (m.group(1), m.group(2) or "0000") if m else ("", "")


def dated_workbooks(root):
    return sorted((p for p in root.glob(WORKBOOK_GLOB) if WORKBOOK_RE.match(p.name)),
                  key=workbook_key)


def newest_workbook(root):
    found = dated_workbooks(root)
    if not found:
        sys.exit(f"no workbook matching 'YYYYMMDD Tibetan_Vocabulary_Dictionary.xlsx' in {root}")
    return found[-1]


def next_version_path(root, src):
    """Today's filename, or a timestamped one if today's is already taken.

    Same-day second versions follow the existing 'YYYYMMDD HHMM …' convention
    from the archive rather than overwriting the day's first version.
    """
    now = datetime.datetime.now()
    plain = root / (now.strftime("%Y%m%d") + " Tibetan_Vocabulary_Dictionary.xlsx")
    if not plain.exists() and plain.resolve() != src.resolve():
        return plain
    return root / (now.strftime("%Y%m%d %H%M") + " Tibetan_Vocabulary_Dictionary.xlsx")


def col_index(letter):
    return ord(letter) - ord("A") + 1


def check_no_slash(value, where):
    """A '/' inside a gloss silently becomes an item boundary on re-import."""
    if "/" in value:
        sys.exit(f"{where}: {value!r} contains '/', which the importer reads as an "
                 "item separator. Rephrase it, or use ';' if you meant two senses.")


def as_date(value):
    """'2026-07-29' -> date(2026, 7, 29). Column U is a real date column."""
    if not value:
        return None
    if isinstance(value, datetime.date):
        return value
    try:
        return datetime.date.fromisoformat(str(value).strip())
    except ValueError:
        sys.exit(f"dateLearned must be YYYY-MM-DD, got {value!r}")


def split_items(value):
    """Inverse of render_pairs for one side: split a paired column on '/'.

    Never on ';' — that separates senses *inside* one item.
    """
    if not value:
        return []
    return [p.strip() for p in str(value).split("/")]


def render_pairs(pairs, where):
    tibs, engs = [], []
    for pair in pairs:
        if len(pair) != 2:
            sys.exit(f"{where}: expected [tibetan, english] pairs, got {pair!r}")
        tib, eng = pair[0].strip(), pair[1].strip()
        check_no_slash(tib, where)
        check_no_slash(eng, where)
        tibs.append(tib)
        engs.append(eng)
    return ITEM_SEP.join(tibs), ITEM_SEP.join(engs)


def read_existing(ws):
    """Tibetan headword -> row number, and the highest sequence per source code."""
    by_tibetan = {}
    max_seq = {}
    for row in range(2, ws.max_row + 1):
        entry_id = ws.cell(row, col_index("A")).value
        if not entry_id:
            continue
        tibetan = ws.cell(row, col_index("B")).value or ""
        by_tibetan.setdefault(tibetan.strip(), row)
        m = re.match(r"([A-Z]+\d*)-(\d+)$", str(entry_id))
        if m:
            code, seq = m.group(1), int(m.group(2))
            max_seq[code] = max(max_seq.get(code, 0), seq)
    return by_tibetan, max_seq


def style_from(ws, template_row, target_row):
    """Give a new row the formatting of the last data row."""
    for letter in FIELDS.values():
        c = col_index(letter)
        src, dst = ws.cell(template_row, c), ws.cell(target_row, c)
        dst._style = copy.copy(src._style)
    if ws.row_dimensions[template_row].height:
        ws.row_dimensions[target_row].height = ws.row_dimensions[template_row].height


def append_to_cell(ws, row, letter, value, sep):
    """Add a source or lesson segment to an existing row without duplicating it."""
    c = ws.cell(row, col_index(letter))
    current = (c.value or "").strip()
    parts = [p.strip() for p in current.split(sep.strip()) if p.strip()] if current else []
    if value.strip() in parts:
        return False
    c.value = (current + sep + value).strip() if current else value
    return True


def apply_new(ws, item, row, source_code, source_name, date_learned, today, max_seq):
    tibetan = item.get("tibetan", "").strip()
    english = item.get("english", "").strip()
    if not tibetan or not english:
        sys.exit(f"new entry needs both tibetan and english: {item!r}")

    seq = max_seq.get(source_code, 0) + 1
    max_seq[source_code] = seq
    entry_id = f"{source_code}-{seq:03d}"

    values = {
        "id": entry_id,
        "tibetan": tibetan,
        "romanization": item.get("romanization", "").strip(),
        "wylie": item.get("wylie", "").strip(),
        "english": english,
        "category": item.get("category", "").strip(),
        "verbStems": item.get("verbStems", "").strip(),
        "volitionality": item.get("volitionality", "").strip(),
        "register": item.get("register", "").strip(),
        "sanskrit": item.get("sanskrit", "").strip(),
        "exampleTibetan": item.get("exampleTibetan", "").strip(),
        "exampleEnglish": item.get("exampleEnglish", "").strip(),
        "notes": item.get("notes", "").strip(),
        "source": item.get("source", source_name).strip(),
        "lesson": item.get("lesson", "").strip(),
        "mastery": "",  # the learner's column; never filled by a script
        # Column U holds real dates in every existing row and column V holds
        # strings. Match both, or Excel left-aligns the new cells as text.
        "dateLearned": as_date(item.get("dateLearned", date_learned)),
        "dateAdded": today,
        "pronCheck": item.get("pronCheck", "derived – reviewed").strip(),
    }
    if values["pronCheck"] not in VALID_PRON:
        sys.exit(f"{entry_id}: pronCheck must be one of {sorted(VALID_PRON)}, "
                 f"got {values['pronCheck']!r}")

    for key, (tib_field, eng_field) in PAIRED.items():
        tib, eng = render_pairs(item.get(key, []), f"{entry_id} {key}")
        values[tib_field] = tib
        values[eng_field] = eng

    for field, letter in FIELDS.items():
        value = values.get(field, "")
        # Write a truly blank cell rather than an empty string, so the column
        # filters and the importer's "" == unknown convention both behave.
        ws.cell(row, col_index(letter)).value = value if value != "" else None
    return entry_id


def main():
    root = pathlib.Path(__file__).resolve().parent.parent
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--entries", required=True, help="JSON list of additions")
    ap.add_argument("--source-code", help="ID prefix for new rows, e.g. NGO7")
    ap.add_argument("--source-name", help="column R value for new rows")
    ap.add_argument("--source-short", default=None,
                    help="compact label for the Look up scope chip, e.g. 'Guru Yoga'")
    ap.add_argument("--date-learned", default=None,
                    help="column U for new rows (default: today)")
    ap.add_argument("--workbook", default=None, help="default: newest in repo root")
    ap.add_argument("--out", default=None, help="default: today's dated filename in repo root")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    import openpyxl

    today = datetime.date.today().isoformat()
    src = pathlib.Path(args.workbook) if args.workbook else newest_workbook(root)
    out = pathlib.Path(args.out) if args.out else next_version_path(root, src)

    if out.resolve() == src.resolve():
        sys.exit(f"refusing to overwrite the source workbook {src.name}. Pass --out.")

    with open(args.entries, encoding="utf-8") as fh:
        items = json.load(fh)
    if not isinstance(items, list) or not items:
        sys.exit("--entries must be a non-empty JSON list")

    wb = openpyxl.load_workbook(src)
    ws = wb[SHEET]
    by_tibetan, max_seq = read_existing(ws)
    template_row = ws.max_row
    next_row = ws.max_row + 1

    added, appended, skipped = [], [], []
    for item in items:
        op = item.get("op", "new")
        tibetan = item.get("tibetan", "").strip()

        if op == "append" or (op == "new" and tibetan in by_tibetan):
            row = by_tibetan.get(tibetan)
            if not row:
                sys.exit(f"append: {tibetan!r} is not in the sheet — use op 'new'")
            if op == "new":
                # Caught rather than trusted: the sheet is one row per word.
                skipped.append((tibetan, row))
            source = item.get("source", args.source_name)
            lesson = item.get("lesson", "").strip()
            if not source or not lesson:
                sys.exit(f"append for {tibetan!r} needs a source and a lesson")
            changed_s = append_to_cell(ws, row, "R", source, SOURCE_SEP)
            changed_l = append_to_cell(ws, row, "S", lesson, LESSON_SEP)
            if changed_s or changed_l:
                appended.append((tibetan, row, lesson))
            continue

        if not args.source_code or not args.source_name:
            sys.exit("new entries need --source-code and --source-name")
        style_from(ws, template_row, next_row)
        entry_id = apply_new(ws, item, next_row, args.source_code, args.source_name,
                             args.date_learned or today, today, max_seq)
        added.append((entry_id, tibetan))
        by_tibetan[tibetan] = next_row
        next_row += 1

    last = next_row - 1
    ws.auto_filter.ref = f"A1:W{last}"
    ws.freeze_panes = FREEZE_PANES

    if "About" in wb.sheetnames:
        about = wb["About"]
        for r in range(1, about.max_row + 1):
            key = about.cell(r, 1).value
            if key == "Entries":
                about.cell(r, 2).value = last - 1
            elif key == "Built":
                about.cell(r, 2).value = datetime.date.today().isoformat()

    registered = None
    if added:
        registered = register_source(root, args.source_code, args.source_name,
                                     args.source_short, args.dry_run)

    print(f"source workbook : {src.name}")
    if registered:
        print(f"registered      : {args.source_code} = {args.source_name!r} "
              f"(chip {registered!r}) in tools/source_names.json")
    print(f"new rows        : {len(added)}")
    for entry_id, tib in added:
        print(f"    {entry_id}  {tib}")
    print(f"sources appended: {len(appended)}")
    for tib, row, lesson in appended:
        print(f"    row {row}  {tib}  += {lesson!r}")
    if skipped:
        print(f"!! {len(skipped)} entry marked 'new' already existed and was "
              "appended to instead (one row per word):")
        for tib, row in skipped:
            print(f"    row {row}  {tib}")
    print(f"total rows      : {last - 1}")

    if args.dry_run:
        print("\ndry run — nothing written")
        return
    wb.save(out)
    print(f"\nwrote {out.name} ({out.stat().st_size:,} bytes)")
    print(f"source workbook {src.name} is unchanged")
    print("\nnext: python3 tools/import_dictionary.py --xlsx " + json.dumps(str(out)))


if __name__ == "__main__":
    main()
