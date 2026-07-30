#!/usr/bin/env python3
"""Correct existing rows in the vocabulary workbook, writing a NEW dated version.

The companion to add_entries.py: that one adds words, this one fixes them. Same
guarantees — never edits a workbook in place, never touches `entry.id`, so review
progress in localStorage survives every correction.

    python3 tools/edit_entries.py --edits corrections.json --dry-run
    python3 tools/edit_entries.py --edits corrections.json

Requires openpyxl. Shared helpers come from add_entries.py so there is one
definition of the column map and the separator rules.

WHY THIS EXISTS RATHER THAN EDITING THE SHEET BY HAND
-----------------------------------------------------
A gloss is written down in more than one place. `དེ་` carries its own English in
column E, and *every other row* that lists དེ་ as a related word or a false
friend repeats that gloss in its own column O or Q. Fixing column E by hand
leaves the copies stale and silently inconsistent — the Look up chips would still
show the wrong meaning. `regloss` fixes all of them at once.

OPERATIONS — a JSON list.

  {"op": "set", "id": "TIB1-019", "english": "that"}

      Change one or more columns on one row. Any short field name from
      add_entries.FIELDS is accepted. `related` and `falseFriends` take pair
      lists and replace the whole column:

      {"op": "set", "id": "TIB1-018",
       "falseFriends": [["དེ་", "that"], ["དེ་རིང་", "today"]]}

  {"op": "regloss", "tibetan": "དེ་", "to": "that", "from": "that; those"}

      Rewrite the gloss for a cross-referenced word everywhere it appears in the
      paired columns of ANY row. `from` is optional; when given, only cells whose
      gloss matches it exactly are touched, which makes the edit auditable and
      refuses to clobber a deliberately different wording.

  {"op": "unlink", "id": "NEC-026", "column": "falseFriends", "tibetan": "དེ་"}

      Remove one item from a paired column, keeping the two sides aligned.

Every operation prints a before/after line. Nothing is written on --dry-run.
"""

import argparse
import json
import pathlib
import sys

sys.path.insert(0, str(pathlib.Path(__file__).resolve().parent))
from add_entries import (  # noqa: E402
    FIELDS, PAIRED, ITEM_SEP, SHEET, as_date, col_index, newest_workbook,
    next_version_path, render_pairs, split_items,
)

DATE_FIELDS = {"dateLearned"}


def read_rows(ws):
    """id -> row number."""
    rows = {}
    for row in range(2, ws.max_row + 1):
        entry_id = ws.cell(row, col_index("A")).value
        if entry_id:
            rows[str(entry_id).strip()] = row
    return rows


def cell_value(ws, row, field):
    return ws.cell(row, col_index(FIELDS[field])).value or ""


def apply_set(ws, rows, edit, log):
    entry_id = edit.get("id")
    row = rows.get(entry_id)
    if not row:
        sys.exit(f"set: no row with id {entry_id!r}")

    for key, value in edit.items():
        if key in ("op", "id"):
            continue

        if key in PAIRED:
            tib_field, eng_field = PAIRED[key]
            before = (cell_value(ws, row, tib_field), cell_value(ws, row, eng_field))
            tib, eng = render_pairs(value, f"{entry_id} {key}")
            ws.cell(row, col_index(FIELDS[tib_field])).value = tib or None
            ws.cell(row, col_index(FIELDS[eng_field])).value = eng or None
            if before != (tib, eng):
                log.append(f"  {entry_id} {key}")
                log.append(f"      - {before[0]!r} || {before[1]!r}")
                log.append(f"      + {tib!r} || {eng!r}")
            continue

        if key not in FIELDS:
            sys.exit(f"set: unknown field {key!r}. Known: {sorted(FIELDS)}")
        if key == "id":
            sys.exit("set: refusing to change an id — that orphans review progress")

        before = cell_value(ws, row, key)
        new = as_date(value) if key in DATE_FIELDS else value
        ws.cell(row, col_index(FIELDS[key])).value = new if new != "" else None
        if str(before) != str(new):
            log.append(f"  {entry_id} {key}")
            log.append(f"      - {before!r}")
            log.append(f"      + {new!r}")


def apply_regloss(ws, rows, edit, log):
    """Rewrite one word's gloss everywhere it is cross-referenced."""
    target = edit["tibetan"].strip()
    to = edit["to"]
    expect = edit.get("from")
    hits = 0
    skipped = []

    for entry_id, row in rows.items():
        for key, (tib_field, eng_field) in PAIRED.items():
            tibs = split_items(cell_value(ws, row, tib_field))
            engs = split_items(cell_value(ws, row, eng_field))
            if not tibs:
                continue
            changed = False
            for i, tib in enumerate(tibs):
                if tib != target or i >= len(engs):
                    continue
                if expect is not None and engs[i] != expect:
                    skipped.append((entry_id, key, engs[i]))
                    continue
                if engs[i] == to:
                    continue
                log.append(f"  {entry_id} {key} -> {target}")
                log.append(f"      - {engs[i]!r}")
                log.append(f"      + {to!r}")
                engs[i] = to
                changed = True
                hits += 1
            if changed:
                ws.cell(row, col_index(FIELDS[eng_field])).value = ITEM_SEP.join(engs)

    if skipped:
        log.append(f"  !! {len(skipped)} cross-reference(s) to {target} left alone "
                   "because the existing gloss did not match 'from':")
        for entry_id, key, actual in skipped:
            log.append(f"      {entry_id} {key}: {actual!r}")
    if not hits and not skipped:
        log.append(f"  (no cross-reference to {target} needed changing)")


def apply_unlink(ws, rows, edit, log):
    entry_id, column = edit["id"], edit["column"]
    row = rows.get(entry_id)
    if not row:
        sys.exit(f"unlink: no row with id {entry_id!r}")
    if column not in PAIRED:
        sys.exit(f"unlink: column must be one of {sorted(PAIRED)}")
    tib_field, eng_field = PAIRED[column]
    tibs = split_items(cell_value(ws, row, tib_field))
    engs = split_items(cell_value(ws, row, eng_field))
    target = edit["tibetan"].strip()
    keep = [(t, engs[i] if i < len(engs) else "") for i, t in enumerate(tibs) if t != target]
    if len(keep) == len(tibs):
        log.append(f"  (no {target} in {entry_id} {column})")
        return
    log.append(f"  {entry_id} {column}: removed {target!r}")
    ws.cell(row, col_index(FIELDS[tib_field])).value = ITEM_SEP.join(t for t, _ in keep) or None
    ws.cell(row, col_index(FIELDS[eng_field])).value = ITEM_SEP.join(e for _, e in keep) or None


HANDLERS = {"set": apply_set, "regloss": apply_regloss, "unlink": apply_unlink}


def main():
    root = pathlib.Path(__file__).resolve().parent.parent
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--edits", required=True, help="JSON list of operations")
    ap.add_argument("--workbook", default=None, help="default: newest in repo root")
    ap.add_argument("--out", default=None, help="default: next dated filename")
    ap.add_argument("--dry-run", action="store_true")
    args = ap.parse_args()

    import openpyxl

    src = pathlib.Path(args.workbook) if args.workbook else newest_workbook(root)
    out = pathlib.Path(args.out) if args.out else next_version_path(root, src)
    if out.resolve() == src.resolve():
        sys.exit(f"refusing to overwrite the source workbook {src.name}. Pass --out.")

    with open(args.edits, encoding="utf-8") as fh:
        edits = json.load(fh)
    if not isinstance(edits, list) or not edits:
        sys.exit("--edits must be a non-empty JSON list")

    wb = openpyxl.load_workbook(src)
    ws = wb[SHEET]
    rows = read_rows(ws)

    log = []
    for edit in edits:
        op = edit.get("op")
        if op not in HANDLERS:
            sys.exit(f"unknown op {op!r}. Known: {sorted(HANDLERS)}")
        HANDLERS[op](ws, rows, edit, log)

    print(f"source workbook : {src.name}")
    print(f"operations      : {len(edits)}")
    print("\nchanges:")
    print("\n".join(log) if log else "  (none)")

    if args.dry_run:
        print("\ndry run — nothing written")
        return
    wb.save(out)
    print(f"\nwrote {out.name} ({out.stat().st_size:,} bytes)")
    print(f"source workbook {src.name} is unchanged")
    print("\nnext: python3 tools/import_dictionary.py")


if __name__ == "__main__":
    main()
