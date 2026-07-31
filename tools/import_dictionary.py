#!/usr/bin/env python3
"""Build tibetan_trainer_data.json from the Tibetan_Vocabulary_Dictionary workbook.

Repeatable: run it again against a newer workbook and it produces the same shape
with more entries. The spreadsheet is content-authoritative; the app is
progress-authoritative (Leitner boxes live in localStorage keyed on entry.id, and
this script never emits progress).

    python3 tools/import_dictionary.py --xlsx "…/20260729 Tibetan_Vocabulary_Dictionary.xlsx"

Requires openpyxl (pip3 install openpyxl).

UNICODE: no normalization anywhere. Tibetan stacks are sequences of separate
codepoints (subjoined U+0F90–U+0FBC), and NFD/NFKC/NFKD all mutate this data —
verified against the workbook. Every string goes from cell to JSON untouched, and
--verify re-reads the output and asserts it round-trips codepoint-for-codepoint.

COLUMN MAP (all 23 columns A–W are used)
    A ID                     -> id, sourceCode (prefix)
    B Tibetan                -> tibetan
    C Romanized Lhasa Pron.  -> romanization
    D Wylie                  -> wylie          (search index only; not displayed)
    E English meaning        -> english, senses (split on ';' — homographs keep
                                every sense, nothing truncated)
    F Word category          -> category
    G Verb stems             -> verbStems
    H Volitional/Transitivity-> volitionality
    I Register               -> register
    J Sanskrit               -> sanskrit
    K Example sentence       -> example.tibetan, example.tokens, example.blank
    L Sentence translation   -> example.english
    M Grammatical notes      -> notes
    N Related words (Tib)  \ -> related[]      positional pairs, split on '/'
    O Related words (Eng)  /
    P False Friends (Tib)  \ -> falseFriends[] positional pairs, split on '/'
    Q False Friends (Eng)  /
    R Source               \ -> sources[], sourceLessons[]  source[n] <-> lesson[n]
    S Lesson / Slide       /    (R split on ';', S split on '|')
    T Mastery                -> mastery        (empty in the workbook -> null)
    U Date learned           -> dateLearned
    V Date added             -> dateAdded
    W Pron. check            -> pronunciationProvenance, pronunciationVerified

SEPARATORS — these were read off the data, not assumed:
    '/'  separates ITEMS in the paired columns N/O and P/Q
    ';'  separates SENSES INSIDE one item ("body; to remain" is one gloss) and
         separates sources in R
    '|'  separates lesson segments in S
So ';' must never be used to split N/O/P/Q: that is what would truncate a
homograph like ལས་ "action; deed; karma; also ablative 'from'".
"""

import argparse
import collections
import datetime
import json
import pathlib
import re
import sys
import unicodedata

COLS = "ABCDEFGHIJKLMNOPQRSTUVW"
SHEET = "Dictionary"

# Human-readable source names, keyed by the ID prefix in column A. Column R names
# the *book* (all three Ngöndro files say "Kagyü Ngöndro (Chariot…)"), so the
# per-code names live in tools/source_names.json, which add_entries.py maintains.
# A code with no entry there is a hard error rather than a raw code leaking into
# the UI.
SOURCE_NAMES_FILE = pathlib.Path(__file__).resolve().parent / "source_names.json"

# Credits and licence notice for the app's About panel. Kept in its own file so
# it can be edited without touching code, and copied into meta so the app still
# reads only tibetan_trainer_data.json.
ATTRIBUTION_FILE = pathlib.Path(__file__).resolve().parent / "attribution.json"


def load_source_names():
    with SOURCE_NAMES_FILE.open(encoding="utf-8") as fh:
        return json.load(fh)["sources"]


def load_attribution():
    if not ATTRIBUTION_FILE.exists():
        return None
    with ATTRIBUTION_FILE.open(encoding="utf-8") as fh:
        doc = json.load(fh)
    return {k: v for k, v in doc.items() if not k.startswith("_")}

# Editorial guidance carried in meta for the app's benefit. Not spreadsheet
# content; the romanization convention restates the About sheet's note.
META_PROSE = {
    "title": "Thomas Hesse — Tibetan vocabulary",
    "romanizationConvention": (
        "Tergar liturgy style: ཟ→Z, ཞ→ZH, བྱ→J. Approximate Lhasa reading, not "
        "phonetic; no tone or vowel length marked."
    ),
    "idIsStable": (
        "The 'id' field is a permanent key. Store review progress against it; "
        "re-importing a newer export must not lose progress."
    ),
    "cardGuidance": {
        "answerSide": (
            "ALWAYS show the romanization when the answer is revealed. There is no "
            "test left to protect at that point, and the pronunciation is part of "
            "what is being learned."
        ),
        "promptSide": (
            "On discrimination cards, hide the romanization only for groups where "
            "hideRomanizationOnPrompt is true — those are groups whose members are "
            "told apart by their romanization, so showing it lets the learner bypass "
            "the script. Where all members share one romanization it is safe, and "
            "useful, to show."
        ),
        "distractorCap": (
            "Some groups are large. Cap wrong answers at three or four, chosen at "
            "random from the group, rather than offering every member."
        ),
        "otherCardTypes": (
            "For plain recall and cloze there is no reason to hide the romanization "
            "anywhere."
        ),
    },
}

UNVERIFIED_PROVENANCE = "derived – reviewed"

# A lesson segment naming a Ngöndro file belongs to the Chariot, which is one
# source in column R. See resolve_source_lessons.
NGONDRO_LESSON = re.compile(r"^Ngöndro\b")
CHARIOT_SOURCE = "Chariot"


# ---------------------------------------------------------------- text helpers

def cell(value):
    """Cell -> string, with no normalization and no reformatting of Tibetan."""
    if value is None:
        return ""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    if isinstance(value, str):
        return value.strip()
    return str(value)


def split_items(value):
    """Split a paired column on '/'. Never on ';' — see the module docstring."""
    if not value:
        return []
    return [p.strip() for p in value.split("/")]


def tokenize(sentence):
    """Split Tibetan on the tsheg, keeping it as each token's terminator.

    Joining the result with "" reproduces the input exactly, which is what the
    cloze renderer relies on.
    """
    return re.findall(r"[^་]*་|[^་]+$", sentence)


def token_base(token):
    """A token stripped of clause punctuation, for headword matching.

    A sentence-final headword wears a shad instead of a tsheg (ཡིན་ -> ཡིན།) and a
    token following a shad carries it as a prefix (། གོང་), so neither compares
    equal to the dictionary form without this.
    """
    return token.strip().lstrip("།༎").strip().rstrip("་།༎").strip()


def locate_blank(headword, tokens):
    """Find the headword's token span in the example sentence.

    Returns {start, length, answer} or None. `answer` is the sentence's own
    wording for the span, not the dictionary form.
    """
    target = [token_base(t) for t in tokenize(headword)]
    if not target or not any(target):
        return None
    n = len(target)
    for i in range(len(tokens) - n + 1):
        if [token_base(t) for t in tokens[i:i + n]] == target:
            return {"start": i, "length": n, "answer": "".join(tokens[i:i + n])}
    return None


# ------------------------------------------------------------------ row -> row

def read_rows(path):
    import openpyxl

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[SHEET]
    header = [cell(c.value) for c in ws[1]]
    if len(header) < len(COLS):
        sys.exit(f"expected {len(COLS)} columns A–W, found {len(header)}")
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        row = {c: cell(raw[i]) for i, c in enumerate(COLS)}
        if not row["A"]:
            continue  # skip blank trailing rows
        rows.append(row)
    about = {}
    if "About" in wb.sheetnames:
        for k, v in wb["About"].iter_rows(values_only=True):
            if k:
                about[cell(k)] = cell(v)
    return rows, about


def pair_columns(row, tib_col, eng_col, report, kind):
    """N/O and P/Q read across as positional pairs."""
    tibs = split_items(row[tib_col])
    engs = split_items(row[eng_col])
    if len(tibs) != len(engs):
        report.pair_mismatch.append((row["A"], kind, row[tib_col], row[eng_col]))
    out = []
    for i, t in enumerate(tibs):
        if not t:
            continue
        # Pair positionally; an item with no counterpart gets "" (= unknown,
        # the model's convention) rather than being dropped or guessed at.
        out.append({"tibetan": t, "english": engs[i] if i < len(engs) else ""})
    return out


def resolve_source_lessons(row, report):
    """Pair column R (sources, ';') with column S (lessons, '|') positionally.

    Some rows carry more lesson segments than sources, because a word that turned
    up again in a later part of an already-listed book gains a lesson segment
    without gaining a source. A surplus "Ngöndro …" segment therefore belongs to
    the Chariot source already listed on the row. Where that source is *not*
    listed, the segment is kept with source: null and sourceMissing: true — the
    data is preserved and the gap is visible, and nothing is invented.

    Returns (sources, pairs) where pairs is one {source, lesson} per lesson
    segment, so the two columns still read across.
    """
    sources = [s.strip() for s in row["R"].split(";") if s.strip()]
    lessons = [l.strip() for l in row["S"].split("|") if l.strip()]

    pairs = [
        {"source": sources[i], "lesson": lesson}
        for i, lesson in enumerate(lessons[:len(sources)])
    ]

    for lesson in lessons[len(sources):]:
        owner = None
        if NGONDRO_LESSON.match(lesson):
            owner = next((s for s in sources if CHARIOT_SOURCE in s), None)
        if owner:
            pairs.append({"source": owner, "lesson": lesson})
            report.lesson_reattached.append((row["A"], lesson, owner))
        else:
            pairs.append({"source": None, "lesson": lesson, "sourceMissing": True})
            report.lesson_unsourced.append((row["A"], lesson))

    if len(lessons) < len(sources):
        report.lesson_short.append((row["A"], row["R"], row["S"]))

    return sources, pairs


def build_entry(row, report):
    entry_id = row["A"]
    match = re.match(r"([A-Z]+\d*)-\d+$", entry_id)
    if not match:
        sys.exit(f"unparseable ID {entry_id!r} — expected e.g. TIB1-001")
    source_code = match.group(1)

    sources, source_lessons = resolve_source_lessons(row, report)

    example = None
    if row["K"]:
        tokens = tokenize(row["K"])
        example = {
            "tibetan": row["K"],
            "english": row["L"],
            "tokens": tokens,
            "blank": locate_blank(row["B"], tokens),
        }

    provenance = row["W"]
    return {
        "id": entry_id,
        "tibetan": row["B"],
        "romanization": row["C"],
        "wylie": row["D"],
        "english": row["E"],
        # Homographs carry several senses in one cell; keep every one, and keep
        # the full string in `english` too so nothing is lost either way.
        "senses": [s.strip() for s in row["E"].split(";") if s.strip()],
        "category": row["F"],
        "verbStems": row["G"],
        "volitionality": row["H"],
        "register": row["I"],
        "sanskrit": row["J"],
        "notes": row["M"],
        "related": pair_columns(row, "N", "O", report, "related"),
        "falseFriends": pair_columns(row, "P", "Q", report, "falseFriends"),
        "sourceCode": source_code,
        "sources": sources,
        "lesson": row["S"],
        "sourceLessons": source_lessons,
        "dateLearned": row["U"] or None,
        "dateAdded": row["V"] or None,
        # Column T is the learner's own column and is empty; progress lives in
        # localStorage keyed on id. Carried through if it is ever filled in.
        "mastery": row["T"] or None,
        "pronunciationVerified": provenance != UNVERIFIED_PROVENANCE,
        "pronunciationProvenance": provenance,
        "example": example,
        "falseFriendGroup": None,  # filled in by build_groups
    }


# ------------------------------------------------------------ false friends

def build_groups(entries):
    """Connected components over the false-friend relation.

    An entry's falseFriends name words in Tibetan; those that are themselves
    entries become edges. The relation is treated as undirected and transitive,
    so a word listed against two different partners lands in one group with all
    of them.
    """
    by_tibetan = collections.defaultdict(list)
    for e in entries:
        by_tibetan[e["tibetan"]].append(e["id"])

    parent = {e["id"]: e["id"] for e in entries}

    def find(x):
        while parent[x] != x:
            parent[x] = parent[parent[x]]
            x = parent[x]
        return x

    def union(a, b):
        ra, rb = find(a), find(b)
        if ra != rb:
            parent[rb] = ra

    for e in entries:
        for ff in e["falseFriends"]:
            for other in by_tibetan.get(ff["tibetan"], []):
                union(e["id"], other)

    members = collections.defaultdict(list)
    for eid in parent:
        members[find(eid)].append(eid)

    by_id = {e["id"]: e for e in entries}
    groups = []
    for ids in sorted((sorted(v) for v in members.values() if len(v) > 1),
                      key=lambda ids: ids[0]):
        romanizations = [by_id[i]["romanization"] for i in ids]
        identical = len(set(romanizations)) == 1
        if identical:
            note = (
                f"All members are romanized {romanizations[0]}. Showing the "
                "pronunciation is harmless here — it cannot disambiguate — and is "
                "worth showing, because it demonstrates that pronunciation will not "
                "save you and the script must be read."
            )
        else:
            note = (
                "Members have different romanizations "
                f"({', '.join(sorted(set(romanizations)))}). Do NOT show the "
                "pronunciation on the QUESTION side: it would let the answer be "
                "picked from the romanization without reading the script. Always "
                "show it on the answer side."
            )
        groups.append({
            "groupId": len(groups) + 1,
            "members": ids,
            "romanizations": romanizations,
            "romanizationsIdentical": identical,
            # The leak runs the opposite way from the obvious guess: risk exists
            # when members DIFFER, because then the romanization alone answers the
            # card. Where all members share one, showing it teaches the real
            # lesson — pronunciation will not save you.
            "hideRomanizationOnPrompt": not identical,
            "note": note,
        })

    for g in groups:
        for mid in g["members"]:
            by_id[mid]["falseFriendGroup"] = g["groupId"]
    return groups


def build_cloze_cards(entries):
    """One card per entry whose headword was locatable in its example sentence."""
    cards = []
    for e in entries:
        ex = e["example"]
        if not ex or not ex["blank"]:
            continue
        cards.append({
            "cardId": "cloze-" + e["id"],
            "entryId": e["id"],
            "tibetan": ex["tibetan"],
            "english": ex["english"],
            "tokens": ex["tokens"],
            "blank": ex["blank"],
        })
    return cards


def build_meta(entries, about, workbook_name):
    codes = []
    for e in entries:
        if e["sourceCode"] not in codes:
            codes.append(e["sourceCode"])

    names = load_source_names()
    unknown = [c for c in codes if c not in names]
    if unknown:
        sys.exit(
            f"no display name for source code(s) {unknown}. Add them to "
            f"{SOURCE_NAMES_FILE.name} (add_entries.py normally does this for you) "
            "so the app does not show a bare code."
        )

    sources = []
    for code in codes:
        dates = [e["dateLearned"] for e in entries
                 if e["sourceCode"] == code and e["dateLearned"]]
        sources.append({
            "code": code,
            "name": names[code]["name"],
            # Compact label for the Look up scope chips; see source_names.json.
            "short": names[code].get("short") or names[code]["name"],
            # Most common date learned across the code's entries; null when the
            # code was never taught on a date (contrast entries).
            "dateLearned": collections.Counter(dates).most_common(1)[0][0] if dates else None,
            "entryCount": sum(1 for e in entries if e["sourceCode"] == code),
        })

    return {
        "schemaVersion": 2,
        "generated": about.get("Built") or datetime.date.today().isoformat(),
        "sourceWorkbook": workbook_name,
        "title": META_PROSE["title"],
        "entryCount": len(entries),
        "falseFriendGroupCount": 0,   # set by build()
        "clozeCardCount": 0,          # set by build()
        "entriesWithFalseFriends": sum(1 for e in entries if e["falseFriends"]),
        "unverifiedPronunciationCount": sum(1 for e in entries
                                            if not e["pronunciationVerified"]),
        "romanizationConvention": META_PROSE["romanizationConvention"],
        "idIsStable": META_PROSE["idIsStable"],
        "cardGuidance": META_PROSE["cardGuidance"],
        "attribution": load_attribution(),
        "sources": sources,
    }


# ---------------------------------------------------------------------- report

class Report:
    def __init__(self):
        self.pair_mismatch = []
        self.lesson_reattached = []
        self.lesson_unsourced = []
        self.lesson_short = []

    def emit(self, entries, groups, cards, previous_ids):
        out = []
        w = out.append
        w(f"entries {len(entries)} · false-friend groups {len(groups)} · cloze cards {len(cards)}")

        if previous_ids is not None:
            new = [e["id"] for e in entries if e["id"] not in previous_ids]
            gone = sorted(previous_ids - {e["id"] for e in entries})
            w(f"ids: +{len(new)} new, {len(gone)} dropped")
            if new:
                counts = collections.Counter(i.split("-")[0] for i in new)
                w("  new: " + ", ".join(f"{k} ×{v}" for k, v in counts.items()))
            if gone:
                # An id disappearing orphans that word's review history.
                w("  !! DROPPED (progress for these becomes orphaned): " + ", ".join(gone))

        no_example = [e["id"] for e in entries if not e["example"]]
        no_blank = [e["id"] for e in entries
                    if e["example"] and not e["example"]["blank"]]
        w(f"examples: {len(entries) - len(no_example)} present, {len(no_example)} absent; "
          f"headword not locatable in {len(no_blank)} of them (no cloze card)")

        if self.lesson_reattached:
            w(f"\ncolumn R/S — {len(self.lesson_reattached)} surplus lesson segment(s) "
              "re-attached to a source already listed on the row:")
            for eid, lesson, owner in self.lesson_reattached:
                w(f"  {eid}: {lesson!r} -> {owner!r}")
        if self.lesson_unsourced:
            w(f"\ncolumn R/S — {len(self.lesson_unsourced)} lesson segment(s) whose source "
              "is NOT listed in column R (kept with source: null, sourceMissing: true):")
            for eid, lesson in self.lesson_unsourced:
                w(f"  {eid}: {lesson!r}")
        if self.lesson_short:
            w(f"\ncolumn R/S — {len(self.lesson_short)} row(s) with fewer lessons than sources:")
            for eid, r, s in self.lesson_short:
                w(f"  {eid}: R={r!r} S={s!r}")
        if self.pair_mismatch:
            w(f"\npaired columns — {len(self.pair_mismatch)} row(s) where the two sides "
              "differ in length (paired positionally, surplus gets english: \"\"):")
            for eid, kind, tib, eng in self.pair_mismatch:
                w(f"  {eid} ({kind}): {tib!r} || {eng!r}")
        return "\n".join(out)


# ----------------------------------------------------------------------- build

def build(xlsx, report):
    rows, about = read_rows(xlsx)
    entries = [build_entry(r, report) for r in rows]

    ids = collections.Counter(e["id"] for e in entries)
    dupes = [i for i, n in ids.items() if n > 1]
    if dupes:
        sys.exit(f"duplicate ids: {dupes}")

    groups = build_groups(entries)
    cards = build_cloze_cards(entries)
    meta = build_meta(entries, about, pathlib.Path(xlsx).name)
    meta["falseFriendGroupCount"] = len(groups)
    meta["clozeCardCount"] = len(cards)

    return {
        "meta": meta,
        "entries": entries,
        "falseFriendGroups": groups,
        "clozeCards": cards,
    }, rows


def verify(data, rows):
    """Assert every workbook string reached the JSON codepoint-for-codepoint."""
    problems = []

    serialized = json.dumps(data, ensure_ascii=False)
    for form in ("NFD", "NFKC", "NFKD"):
        if unicodedata.normalize(form, serialized) != serialized:
            break
    else:
        problems.append(
            "expected this data to be sensitive to Unicode normalization; it no "
            "longer is, which may mean it was normalized upstream"
        )

    by_id = {e["id"]: e for e in data["entries"]}
    direct = {"B": "tibetan", "C": "romanization", "D": "wylie", "E": "english",
              "F": "category", "G": "verbStems", "H": "volitionality",
              "I": "register", "J": "sanskrit", "M": "notes", "S": "lesson",
              "W": "pronunciationProvenance"}
    for row in rows:
        e = by_id[row["A"]]
        for col, field in direct.items():
            if e[field] != row[col]:
                problems.append(f"{row['A']} column {col}: {row[col]!r} != {e[field]!r}")
        if row["K"]:
            if e["example"]["tibetan"] != row["K"]:
                problems.append(f"{row['A']} column K altered")
            if "".join(e["example"]["tokens"]) != row["K"]:
                problems.append(f"{row['A']} tokens do not rejoin to column K")
        # Every lesson segment survives, and every source is still named.
        if sorted(p["lesson"] for p in e["sourceLessons"]) != \
           sorted(l.strip() for l in row["S"].split("|") if l.strip()):
            problems.append(f"{row['A']} lost or altered a lesson segment")
        for item, col in ((e["related"], "N"), (e["falseFriends"], "P")):
            if [i["tibetan"] for i in item] != [p for p in split_items(row[col]) if p]:
                problems.append(f"{row['A']} column {col} items altered")
    return problems


def main():
    here = pathlib.Path(__file__).resolve().parent.parent
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--xlsx", default=None,
                    help="path to the workbook (default: newest dated workbook in the repo root)")
    ap.add_argument("--out", default=str(here / "tibetan_trainer_data.json"))
    ap.add_argument("--previous", default=None,
                    help="prior JSON to diff ids against (default: --out if it exists)")
    ap.add_argument("--sample", type=int, default=0,
                    help="print N sample entries instead of writing")
    ap.add_argument("--sample-ids", default="",
                    help="comma-separated ids to print instead of writing")
    ap.add_argument("--dry-run", action="store_true", help="report only, write nothing")
    args = ap.parse_args()

    xlsx = args.xlsx
    if not xlsx:
        # "20260729 …" or the same-day variant "20260729 1930 …". Sorted by parsed
        # date/time, because lexicographically '20260729 1930' < '20260729 T…'.
        pattern = re.compile(r"^(\d{8})(?: (\d{4}))? Tibetan_Vocabulary_Dictionary\.xlsx$")
        found = sorted(
            ((pattern.match(p.name), p) for p in here.glob("20* Tibetan_Vocabulary_Dictionary.xlsx")),
            key=lambda pair: (pair[0].group(1), pair[0].group(2) or "0000") if pair[0] else ("", ""),
        )
        found = [p for m, p in found if m]
        if not found:
            sys.exit("no dated workbook in the repo root — pass --xlsx")
        xlsx = str(found[-1])
        # stderr, so --sample-ids output stays pipeable JSON on stdout.
        print(f"using {pathlib.Path(xlsx).name}", file=sys.stderr)

    report = Report()
    data, rows = build(xlsx, report)

    problems = verify(data, rows)
    if problems:
        print("VERIFY FAILED:", file=sys.stderr)
        for p in problems[:40]:
            print("  " + p, file=sys.stderr)
        sys.exit(1)

    previous_ids = None
    prev_path = pathlib.Path(args.previous or args.out)
    if prev_path.exists():
        with prev_path.open(encoding="utf-8") as fh:
            previous_ids = {e["id"] for e in json.load(fh)["entries"]}

    if args.sample or args.sample_ids:
        by_id = {e["id"]: e for e in data["entries"]}
        chosen = ([by_id[i.strip()] for i in args.sample_ids.split(",") if i.strip()]
                  or data["entries"][:args.sample])
        print(json.dumps(chosen, ensure_ascii=False, indent=2))
        print("\n---\n" + report.emit(data["entries"], data["falseFriendGroups"],
                                      data["clozeCards"], previous_ids), file=sys.stderr)
        return

    print(report.emit(data["entries"], data["falseFriendGroups"],
                      data["clozeCards"], previous_ids))
    if args.dry_run:
        print("\ndry run — nothing written")
        return

    out = pathlib.Path(args.out)
    with out.open("w", encoding="utf-8") as fh:
        json.dump(data, fh, ensure_ascii=False, indent=1)
        fh.write("\n")
    print(f"\nwrote {out} ({out.stat().st_size:,} bytes)")


if __name__ == "__main__":
    main()
